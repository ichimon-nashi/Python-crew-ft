from tkinter import *
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
import datetime
import time

BGCOLOR = "#292828"
FONT_NAME = "Taipei Sans TC Beta"
WHITE = "#FFFFFF"
khh_crew = ['22119','51892', '33107', '34011', '51043', '51837', '53522', '55033', '55120', '56392', '59161',
            '59230', '59262']

# ---------------------------- 檢核 ------------------------------- #
def run_inspection():
    login_id = login_id_entry.get()
    login_pass = password_entry.get()
    check_date = int(inspect_duration_entry.get())
    flight_time_list = []
    inspect_date_list = []
    cell_endpoint = ['b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u',
                     'v', 'w', 'x', 'y', 'z', 'aa', 'ab', 'ac', 'ad', 'ae', 'af', 'ag']
    # 預檢天數
    today = datetime.date.today()
    # 設定time delta(時間間隔)
    day_delta = datetime.timedelta(days=1)

    driver = webdriver.Chrome(r"C:\Users\user\Desktop\Python\chromedriver_win32\chromedriver.exe")
    # URL redacted
    driver.get('webscrape URL')

    # driver.implicitly_wait(10)

    # 登入頁面
    login_ID = driver.find_element(By.ID, "txtUserID")
    login_ID.send_keys(login_id)
    login_password = driver.find_element(By.ID, "txtPassword")
    login_password.send_keys(login_pass)
    login_button = driver.find_element(By.ID, "btnloginstyle001").click()

    # driver.implicitly_wait(10)

    # 組員報到摘要
    crew_fatigue_check = driver.find_element(By.ID, "li_t02_tktrec").click()

    # 檢核日期打勾
    checkbox = driver.find_element(By.ID, "ContentPlaceHolder1_cb_d").click()

    for date in range(check_date):
        new_date = today + date * day_delta
        inspect_id = driver.find_element(By.ID, "ContentPlaceHolder1_TextBox1").get_attribute("value")
        inspect_name = driver.find_element(By.ID, "ContentPlaceHolder1_L_name").get_attribute("textContent")
        inspect_date = driver.find_element(By.ID, "ContentPlaceHolder1_tb_date")
        inspect_date.clear()
        inspect_date.send_keys(f"{new_date}")
        inspect_date_list.append(new_date.strftime("%Y/%m/%d"))
        run_check = driver.find_element(By.ID, "ContentPlaceHolder1_Button1").click()
        crew_ft = driver.find_element(By.ID, "ContentPlaceHolder1_L_ft")
        flight_time_list.append(crew_ft.get_attribute("textContent"))

    driver.close()

    print(inspect_date_list)
    print(flight_time_list)

    headers = ["員編", "姓名"]
    workbook_name = "khh_crew_ft.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"{login_id}飛時"
    ws.append(headers)
    ws['A2'] = inspect_id
    ws['B2'] = inspect_name

    for row in ws['c1':cell_endpoint[check_date] + "1"]:
        for index, cell in enumerate(row):
            cell.value = inspect_date_list[index]

    for row in ws['c2':cell_endpoint[check_date] + "2"]:
        for index, cell in enumerate(row):
            cell.value = flight_time_list[index]

    wb.save(filename=workbook_name)

    window.destroy()
# ---------------------------- UI SETUP ------------------------------- #
window = Tk()
window.title("高雄組員FT預檢")
window.minsize(width=500, height=600)
window.maxsize(width=500, height=600)
window.config(padx=10, pady=5, bg=BGCOLOR)

canvas = Canvas(width=450, height=350, bg=BGCOLOR, highlightthickness=0)
skull_logo = PhotoImage(file="logo.png")
canvas.create_image(250, 170, image=skull_logo, anchor= CENTER)
canvas.grid(column=0, row=0, columnspan=2)

#Labels
title_label = Label(text="組員FT預檢", bg=BGCOLOR, fg=WHITE, font=(FONT_NAME, 40))
title_label.grid(column=1, row=1)
EIP_login_label = Label(text="請登入您的EIP帳號密碼", bg=BGCOLOR, fg=WHITE, font=(FONT_NAME, 18, "italic"))
EIP_login_label.config(pady=5)
EIP_login_label.grid(column=1, row=2)
login_id_label = Label(text="帳號:", bg=BGCOLOR, fg=WHITE, width=5, font=(FONT_NAME, 12))
login_id_label.grid(column=0, row=3)
password_label = Label(text="密碼:", bg=BGCOLOR, fg=WHITE, font=(FONT_NAME, 12))
password_label.grid(column=0, row=4)
inspect_duration_label = Label(text="預檢天數:", bg=BGCOLOR, fg=WHITE,  font=(FONT_NAME, 12))
inspect_duration_label.grid(column=0, row=5)

#Entries
login_id_entry = Entry(width=25)
login_id_entry.grid(column=1, row=3, pady=3)
password_entry = Entry(width=25)
password_entry.config(show="*")
password_entry.grid(column=1, row=4, pady=3)
inspect_duration_entry = Entry(width=25)
inspect_duration_entry.insert(0, "1")
inspect_duration_entry.grid(column=1, row=5, pady=3)

# Buttons
inspect_button = Button(text="預檢GO!", font=(FONT_NAME, 12), command=run_inspection)
inspect_button.config(pady=10)
inspect_button.grid(column=1, row=6, pady=5)

window.mainloop()